import openpyxl


class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expected_response = None
        self.actual_response = None
        self.result = None
        self.sql = None


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row

        cases = []
        for r in range(2, max_row+1):
            case = Case()
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected_response = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value

            cases.append(case)

        return cases

    def write_back(self, row, actual_response, result):
        self.sheet.cell(row=row, column=7).value = actual_response
        self.sheet.cell(row=row, column=8).value = result
        self.workbook.save(self.file_name)
        self.workbook.close()

